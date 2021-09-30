#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl


# 操作excel
class DoExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        '''获取需要打开的excel文件和操作的表单'''
        try:
            wb = openpyxl.load_workbook(self.file_name)
            sheet = wb[self.sheet_name]
        except Exception as e:
            print("excel文件或sheet不存在：{0}".format(e))
            raise e

        # 存储从 excel里面读取的数据
        test_data = []

        for i in range(1, sheet.max_row + 1):
            get_test_data = {}
            # 将excel数据一个个读取出来，放到字典里面
            get_test_data['col1'] = sheet.cell(i, 1).value
            get_test_data['col2'] = sheet.cell(i, 2).value
            get_test_data['col3'] = sheet.cell(i, 3).value
            get_test_data['col4'] = sheet.cell(i, 4).value
            get_test_data['col5'] = sheet.cell(i, 5).value
            get_test_data['col6'] = sheet.cell(i, 6).value

            test_data.append(get_test_data)

        return test_data
        # print(get_data)
        # print(test_data)

    def write_back(self, file_name, sheet_name, row, value4):
        # 打开excel，定位到表单
        wb = openpyxl.load_workbook(file_name)
        sheet = wb[sheet_name]
        # sheet.cell(row, 1).value = value1
        sheet.cell(row, 4).value = value4
        wb.save(file_name)


if __name__ == '__main__':

    do_excel = DoExcel(r'D:\测试数据\标准诊断-diagnose-standard2.xlsx', 'result')
    excel = do_excel.get_data()

    print(len(excel))
    # print(excel)
    z = 1
    for x in range(1, len(excel)):
        col1 = excel[x]['col1']
        col4 = excel[x]['col4']
        # print(col1, col4)
        # print(type(col1),type(col4))
        if col1 == col4:
            print(x)
            z = z + 1
            do_excel.write_back(r'D:\测试数据\标准诊断-diagnose-standard2.xlsx', 'result', x, '')
            print(col1, col4)
    # print(z)
