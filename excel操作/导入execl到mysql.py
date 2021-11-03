#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl
from excel操作.dbMyql_util import MysqlDb

# 操作excel
class DoExcel:

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.mysql = MysqlDb()

    def get_data(self):
        '''获取需要打开的excel文件和操作的表单'''
        try :
            wb = openpyxl.load_workbook(self.file_name)
            sheet = wb[self.sheet_name]
        except Exception as e:
            print("excel文件或sheet不存在：{0}".format(e))
            raise e

        table_name = "table_4"
        self.create_table(table_name)
        # 存储从 excel里面读取的数据
        test_data = []
        for i in range(1,sheet.max_row+1):
            get_test_data = {}
            # 将excel数据一个个读取出来，放到字典里面
            col_1 = get_test_data['col1'] = sheet.cell(i,1).value
            col_2 =get_test_data['col2'] = sheet.cell(i, 2).value
            col_3 =get_test_data['col3'] = sheet.cell(i, 3).value
            col_4 =get_test_data['col4'] = sheet.cell(i, 4).value
            col_5 =get_test_data['col5'] = sheet.cell(i, 5).value
            col_6 =get_test_data['col6'] = sheet.cell(i, 6).value
            col_7 =get_test_data['col7'] = sheet.cell(i, 7).value
            col_8 =get_test_data['col8'] = sheet.cell(i, 8).value
            col_9 =get_test_data['col9'] = sheet.cell(i, 9).value
            col_10 =get_test_data['col10'] = sheet.cell(i, 10).value
            col_11 =get_test_data['col11'] = sheet.cell(i, 11).value
            col_12 =get_test_data['col12'] = sheet.cell(i, 12).value

            test_data.append(get_test_data)

            insert_sql = '''INSERT INTO excel_data.{0} (col_1, col_2,col_3, col_4,col_5, col_6,col_7, col_8,col_9, col_10,col_11,col_12) VALUES ('{1}', '{2}', '{3}', '{4}', '{5}', '{6}', '{7}', '{8}', '{9}', '{10}', '{11}', '{12}');
            '''.format(table_name,col_1, col_2,col_3, col_4,col_5, col_6,col_7, col_8,col_9, col_10,col_11,col_12)
            self.mysql.execut(insert_sql)
            print("第{0}条，插入成功".format(i))


        return test_data
        # print(get_data)
        # print(test_data)

    def write_back(self, file_name, sheet_name, row, result, testResult):
        # 打开excel，定位到表单
        wb = openpyxl.load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(row, 8).value = result
        sheet.cell(row, 9).value = testResult
        wb.save(file_name)

    def create_table(self,table_name):

        creat_sql='''create table if not exists {0}
            (
                col_1 varchar(1000) null,
                col_2 varchar(1000) null,
                col_3 varchar(1000) null,
                col_4 varchar(1000) null,
                col_5 varchar(1000) null,
                col_6 varchar(1000) null,
                col_7 varchar(1000) null,
                col_8 varchar(1000) null,
                col_9 varchar(1000) null,
                col_10 varchar(1000) null,
                col_11 varchar(1000) null,
                col_12 varchar(1000) null
            );
    ''' .format(table_name)
        print(table_name)
        self.mysql.execut(creat_sql)
        # sql_res = self.mysql.query('select * from table_1;')
        print(table_name,"表创建成功")

if __name__ == '__main__':

    do_excel = DoExcel(r'D:\python_work\synyi_test\test_sim_zhengce\需要导入的政策指标集.xlsx', 'Sheet1')
    do_excel.get_data()
    # print(res)